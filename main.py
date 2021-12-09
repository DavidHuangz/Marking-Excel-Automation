# pip install pandas
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def csv_to_excel(fileIn, fileOut):
    # 2) read csv file
    csv = pd.read_csv(fileIn)

    # 3) create excel writer
    excelWriter = pd.ExcelWriter(fileOut)

    # 4) convert csv to excel
    csv.to_excel(
        excelWriter,
        index_label='ABC',
        index=False,
        float_format='%.2f',
        # header = False,
        freeze_panes=(3, 1),
        sheet_name='grades from csv'
    )
    # 5) save Excel file
    excelWriter.save()


def xlsx_to_csv(fileOut, fileMarkedOut):
    # 2) read csv file
    data_xls = pd.read_excel(fileOut)
    data_xls.to_csv(fileMarkedOut, encoding='utf-8', index=False)


def ExtractTxt(canvasText, UPI, Score):
    # Step 2 Extract data from txt file
    with open(canvasText) as fp:
        contents = fp.read()
        try:
            for entry in contents.split('\n'):
                UPI.append(int(entry.split(' ')[1]))
                Score.append(int(entry.split(' ')[2]))
        except Exception as e:
            print(e)


# Step 3 Add grades to the Excel sheet
def AutoMarking(wb, ws, row_count, UPI, Score, fileOut, UPIChar, ScoreChar, RowStart):
    # Check if upi matches
    UPI_index = RowStart
    for cols in range(RowStart, row_count+1):
        try:
            if UPI[UPI_index] == ws[UPIChar + str(cols)].value:
                # Add score
                ws[ScoreChar + str(cols)].value = Score[UPI_index]
                # Align left
                # ws[ScoreChar + str(cols)].alignment = Alignment(horizontal='left')
                print('\nUPI: ' + str(UPI[UPI_index]) + ', Score: ' + str(Score[UPI_index]))
                print('Checking UPI: ' + str(UPI[UPI_index]) + ' for ' + UPIChar + str(cols))
                print('Adding' + ' score ' + str(Score[UPI_index]) + ' for ' + ScoreChar + str(cols) + '\n')
                UPI_index += 1

        except Exception as e:
            print(e)
            break

    wb.save(fileOut)
    print('Marking completed for ' + str(UPI_index - RowStart) + ' student(s)')


def main():
    # Constants
    fileOut = 'grades.xlsx'
    fileIn = 'grades.csv'
    fileMarkedOut = 'gradesMarked.csv'
    canvasText = 'canvasMarking.txt'

    # Excel row characters for different assignments and UPI
    UPIChar = 'B'
    ScoreChar = 'H'
    RowStart = 3  # Row where first student is located

    UPI = []
    Score = []
    # 4 Columns are empty so they are None
    for rowX in range(RowStart):
        UPI.append(None)
        Score.append(None)

    # Convert csv to xlsx
    csv_to_excel(fileIn, fileOut)

    # load excel file
    wb = load_workbook(fileOut)
    ws = wb.active

    # Get total rows
    sheet = wb.worksheets[0]
    row_count = sheet.max_row
    print('Total rows: ' + str(row_count))

    # Extract data from txt file
    ExtractTxt(canvasText, UPI, Score)
    print(str(UPI) + '\n' + str(Score))

    #  Add grades to the Excel sheet
    AutoMarking(wb, ws, row_count, UPI, Score, fileOut, UPIChar, ScoreChar, RowStart)

    # Convert back to csv
    xlsx_to_csv(fileOut, fileMarkedOut)


if __name__ == '__main__':
    main()
